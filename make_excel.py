from openpyxl import load_workbook 
from userdata import UserData
import pandas as pd 
from datetime import  datetime 


#check date and name the file 
current_date = datetime.now().date().__str__() 
file_name = 'demo.xlsx'


#load wb and append sheet in case of no existing
wb = load_workbook(filename = file_name) 
ws=wb.active

if current_date in wb.sheetnames: 
    #append data with df 
    df = pd.DataFrame({'Name': ['A', 'B', 'C', 'D'],
                   'Age': [5, 5, 5, 5]})
    df2 = pd.read_excel(file_name, sheet_name=current_date)
    df1 = pd.concat([df2,df])

    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    writer.book=wb
    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

    # Convert the dataframe to an XlsxWriter Excel object.
    df1.to_excel(writer, sheet_name= current_date, index=False)

    # Close the Pandas Excel writer and output the Excel file.
    writer.save()
else: 
    ws = wb.create_sheet(title=current_date)
    wb.save(filename=file_name)
    df = pd.DataFrame({'Name': ['A', 'B', 'C', 'D'],
                   'Age': [0, 0, 0, 0]})
    #add data 
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    writer.book=wb
    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
     # Convert the dataframe to an XlsxWriter Excel object.
    df.to_excel(writer, sheet_name= current_date, index=False)

    # Close the Pandas Excel writer and output the Excel file.
    writer.save()



  
    




